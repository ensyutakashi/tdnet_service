import duckdb
import openpyxl
import pandas as pd
import os
import sys

# --- 設定 ---
DB_FILE = 'TDnet.duckdb'
TABLE_NAME = 'disclosure_info'

# 初期移行用ファイル（マスタ）
MASTER_FILE = 'TDnet適時開示情報.xlsm'
SHEET_NAME = '適時開示情報'

# 日次追加用ファイル（例）
# DAILY_FILE = 'TDNET_Output.xlsx' 

def extract_data_with_links(excel_path, sheet_name=None):
    """
    Excelを開き、データとハイパーリンク(D, E列)を抽出してDataFrameを作成する
    """
    print(f"読み込み中: {excel_path} ...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True) # 数値・計算結果を取得
    except FileNotFoundError:
        print(f"エラー: ファイルが見つかりません -> {excel_path}")
        return None

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"エラー: シート '{sheet_name}' が見つかりません。")
            return None
        ws = wb[sheet_name]
    else:
        ws = wb.active

    data = []
    
    # ヘッダー取得 (1行目)
    headers = [cell.value for cell in ws[1]]
    
    # A-P列 (16列) までが既存データと想定
    # Q, R列を追加
    headers_extended = headers[:16] + ['表題リンク', 'XBRLリンク']
    
    # 2行目からデータ取得
    for row in ws.iter_rows(min_row=2):
        # A-P列の値を取得 (index 0-15)
        row_values = [cell.value for cell in row[:16]]
        
        # D列(index 3) のリンク抽出
        link_d = row[3].hyperlink.target if row[3].hyperlink else None
        
        # E列(index 4) のリンク抽出
        link_e = row[4].hyperlink.target if row[4].hyperlink else None
        
        # 行データ結合
        row_values.append(link_d) # Q列相当
        row_values.append(link_e) # R列相当
        
        data.append(row_values)

    df = pd.DataFrame(data, columns=headers_extended)
    print(f"抽出完了: {len(df)} 件")
    return df

def create_or_append_duckdb(df, db_path, table_name):
    """
    DataFrameをDuckDBに保存（テーブルがなければ作成、あれば追記）
    """
    if df is None or df.empty:
        print("データがないため処理を中断します。")
        return

    con = duckdb.connect(db_path)
    
    # テーブルの存在確認
    tables = con.execute("SHOW TABLES").fetchall()
    table_exists = any(table_name in t for t in tables)

    if not table_exists:
        print(f"DuckDBを作成中: {db_path}")
        con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM df")
        print("初期作成完了。")
    else:
        print(f"DuckDBに追記中: {db_path}")
        # 列の整合性を保つため INSERT INTO SELECT を使用
        con.execute(f"INSERT INTO {table_name} SELECT * FROM df")
        print("追記完了。")
    
    con.close()

# --- 実行ブロック ---
if __name__ == '__main__':
    # 1. マスタファイルが存在する場合、それを元にDB作成（初回のみ推奨）
    if not os.path.exists(DB_FILE):
        print(">>> 初期移行モード (TDnet適時開示情報.xlsm -> DuckDB)")
        df_master = extract_data_with_links(MASTER_FILE, SHEET_NAME)
        create_or_append_duckdb(df_master, DB_FILE, TABLE_NAME)
    else:
        print(f">>> 追記モード ({DB_FILE} は既に存在します)")
        print("日次ファイルを追加するには、コード内の DAILY_FILE を指定して extract_data_with_links を呼び出してください。")
        # 例: 日次ファイルの処理
        # df_daily = extract_data_with_links('TDNET_Output.xlsx')
        # create_or_append_duckdb(df_daily, DB_FILE, TABLE_NAME)

    print("\n処理が終了しました。")