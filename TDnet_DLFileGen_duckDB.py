# --- metadata ---
# description:(仮)TDnetから適時開示を取得しExcel出力する
# システム構成図: cnvs_TDnet適時開示情報閲覧サービス.canvas
# --- metadata ---

# TDnetから適時開示情報を取得し、Excelに出力するスクリプト 
# -*- coding: utf-8 -*-
"""
TDnet 適時開示（公開日：期間指定）を統合してExcel出力。
サイト列そのまま（A:G = 時刻, コード, 会社名, 表題, XBRL, 上場取引所, 更新履歴）
H列に公開日（yy/mm/dd, 日付型）を追加。
表題・XBRLはハイパーリンク、青白ストライプのテーブル、枠固定あり。

使い方例:
cd "\\LS720D7A9\TakashiBK\投資\MyPython"
    python TDnet_DLFileGen.py --start 20260126 --end 20260127 --out TDNET_Output.xlsx
DL先は　\\LS720D7A9\TakashiBK\投資\MyPython
"""

import argparse
import sys
import time
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import csv

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import duckdb
import os

VERSION = "TDNET Range + H列(公開日) v1.04"

BASE = "https://www.release.tdnet.info/inbs/"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    )
}

# A～H のヘッダー（H=公開日）
COLS = ["時刻", "コード", "会社名", "表題", "XBRL", "上場取引所", "更新履歴", "公開日"]


def fetch_page_html(sess: requests.Session, page_path: str) -> Optional[str]:
    """ページHTML取得（簡易リトライ）。成功時は文字列、失敗時 None。"""
    url = BASE + page_path
    for i in range(3):  # リトライ 3回
        try:
            r = sess.get(url, headers=HEADERS, timeout=20)
            if r.status_code == 200 and "main-list-table" in r.text:
                r.encoding = r.apparent_encoding or "utf-8"
                return r.text
        except requests.RequestException:
            pass
        time.sleep(1 + i)
    return None


def parse_rows(html: str, date_str: str) -> List[Dict]:
    """main-list-table をパースして、行辞書のリストを返す。H列用に公開日を付与。"""
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table", id="main-list-table")
    rows: List[Dict] = []
    if not table:
        return rows

    pub_date = datetime.strptime(date_str, "%Y%m%d").date()  # H列の公開日（日付型）
    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 7:
            continue

        # 0:時刻, 1:コード, 2:会社名
        t_time = tds[0].get_text(strip=True)
        t_code = tds[1].get_text(strip=True)  # 英字混在コードあり→文字列保持
        t_name = tds[2].get_text(strip=True)

        # 3:表題（PDFリンク）
        title_td = tds[3]
        a = title_td.find("a")
        title_txt = title_td.get_text(strip=True)
        pdf_url = BASE + a["href"].lstrip("./") if (a and a.get("href")) else None

        # 4:XBRL（zipリンクがあるときのみ）
        x_td = tds[4]
        xa = x_td.find("a")
        x_url = BASE + xa["href"].lstrip("./") if (xa and xa.get("href")) else None
        x_text = "XBRL" if x_url else ""

        # 5:上場取引所, 6:更新履歴
        place = tds[5].get_text(strip=True)
        hist = tds[6].get_text(strip=True)

        rows.append({
            "時刻": t_time, "コード": t_code, "会社名": t_name,
            "表題": title_txt, "表題URL": pdf_url,
            "XBRL": x_text, "XBRLURL": x_url,
            "上場取引所": place, "更新履歴": hist,
            "公開日": pub_date,
        })
    return rows


def get_latest_db_data(max_retries=3, retry_delay=1):
    """データベースから最新の検索用データを取得（リトライ機能付き）"""
    db_path = r"\\LS720D7A9\TakashiBK\投資\TDNET\TDnet適時情報開示サービス\tdnet.duckdb"
    
    if not os.path.exists(db_path):
        print(f"エラー: データベースファイルが見つかりません: {db_path}")
        return None
    
    con = None
    for attempt in range(1, max_retries + 1):
        try:
            con = duckdb.connect(database=db_path, read_only=True)
            query = """
            SELECT 
                "公開日", 
                "時刻", 
                "コード", 
                "会社名", 
                "表題" 
            FROM disclosure_info 
            ORDER BY "公開日" DESC, "時刻" DESC 
            LIMIT 1
            """
            result = con.execute(query).fetchone()
            con.close()
            
            if result:
                koukai_bi, shikoku, code, company, title = result
                search_data = f"{koukai_bi}&{shikoku}&{code}&{company}&{title}"
                return {
                    'search_data': search_data,
                    'date': koukai_bi,
                    'time': shikoku,
                    'code': code,
                    'company': company,
                    'title': title
                }
            return None
            
        except Exception as e:
            if attempt < max_retries:
                wait_time = retry_delay * attempt
                print(f"リトライ中... ({attempt}/{max_retries}) - {datetime.now().strftime('%H:%M:%S')}")
                print(f"エラー: {str(e)}")
                print(f"{wait_time}秒後に再試行します...")
                time.sleep(wait_time)
            else:
                print(f"データベース接続エラー: {e}")
                return None
        
        finally:
            if con is not None:
                try:
                    con.close()
                except:
                    pass


def get_db_data_by_date(target_date, max_retries=3, retry_delay=1):
    """指定日のデータベースデータを取得（リトライ機能付き）"""
    db_path = r"\\LS720D7A9\TakashiBK\投資\TDNET\TDnet適時情報開示サービス\tdnet.duckdb"
    
    if not os.path.exists(db_path):
        print(f"エラー: データベースファイルが見つかりません: {db_path}")
        return []
    
    con = None
    for attempt in range(1, max_retries + 1):
        try:
            con = duckdb.connect(database=db_path, read_only=True)
            query = """
            SELECT 
                "時刻", 
                "コード", 
                "会社名", 
                "表題" 
            FROM disclosure_info 
            WHERE DATE("公開日") = ?
            ORDER BY "時刻" DESC
            """
            result = con.execute(query, [target_date]).fetchall()
            con.close()
            return result
            
        except Exception as e:
            if attempt < max_retries:
                wait_time = retry_delay * attempt
                print(f"リトライ中... ({attempt}/{max_retries}) - {datetime.now().strftime('%H:%M:%S')}")
                print(f"エラー: {str(e)}")
                print(f"{wait_time}秒後に再試行します...")
                time.sleep(wait_time)
            else:
                print(f"データベース接続エラー: {e}")
                return []
        
        finally:
            if con is not None:
                try:
                    con.close()
                except:
                    pass


def export_diff_to_csv(diff_data, target_date):
    """差分データをCSVファイルに出力"""
    if not diff_data['only_in_web']:
        print("差分データがないため、CSV出力をスキップします")
        return
    
    # CSVファイル名を作成（現在のディレクトリに保存）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    csv_filename = f"TDNET_Diff_{target_date}_{timestamp}.csv"
    csv_path = csv_filename  # 現在のディレクトリに保存
    
    try:
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # ヘッダー
            writer.writerow(['時刻', 'コード', '会社名', '表題', '表題URL', 'XBRL', 'XBRLURL', '上場取引所', '更新履歴', '公開日'])
            
            # 差分データを書き込み
            web_data = scrape_one_day(target_date.replace('-', ''))
            for row in web_data:
                row_key = (row["時刻"], row["コード"], row["会社名"], row["表題"])
                if row_key in diff_data['only_in_web']:
                    writer.writerow([
                        row["時刻"],
                        row["コード"],
                        row["会社名"],
                        row["表題"],
                        row.get("表題URL", ""),
                        row["XBRL"],
                        row.get("XBRLURL", ""),
                        row["上場取引所"],
                        row["更新履歴"],
                        row["公開日"].strftime('%Y-%m-%d')
                    ])
        
        print(f"\n✅ 差分データをCSV出力しました: {csv_path}")
        print(f"   差分件数: {len(diff_data['only_in_web'])}件")
        
    except Exception as e:
        print(f"CSV出力エラー: {e}")


def check_daily_diff(target_date, export_csv=False):
    """指定日のTDNETサイトとデータベースの差分をチェック"""
    print(f"\n=== {target_date} の差分チェック ===")
    
    # データベースから指定日のデータを取得
    db_data = get_db_data_by_date(target_date)
    print(f"データベース件数: {len(db_data)}件")
    
    # TDNETサイトからデータを取得
    date_str = target_date.replace('-', '')
    web_data = scrape_one_day(date_str)
    print(f"TDNETサイト件数: {len(web_data)}件")
    
    # 比較用にデータを整形
    db_set = set()
    for row in db_data:
        db_set.add((row[0], row[1], row[2], row[3]))  # (時刻, コード, 会社名, 表題)
    
    web_set = set()
    for row in web_data:
        web_set.add((row["時刻"], row["コード"], row["会社名"], row["表題"]))
    
    # 差分を計算
    only_in_db = db_set - web_set
    only_in_web = web_set - db_set
    
    print(f"\n--- 差分結果 ---")
    print(f"データベースのみに存在: {len(only_in_db)}件")
    if only_in_db:
        for item in sorted(only_in_db):
            print(f"  DBのみ: {item}")
    
    print(f"\nTDNETサイトのみに存在: {len(only_in_web)}件")
    if only_in_web:
        for item in sorted(only_in_web):
            print(f"  Webのみ: {item}")
    
    if not only_in_db and not only_in_web:
        print("✅ データは一致しています")
    
    result = {
        'db_count': len(db_data),
        'web_count': len(web_data),
        'only_in_db': only_in_db,
        'only_in_web': only_in_web,
        'is_match': len(only_in_db) == 0 and len(only_in_web) == 0
    }
    
    # CSV出力オプション
    if export_csv and only_in_web:
        export_diff_to_csv(result, target_date)
    
    return result


def check_and_download_next_days(target_date, download_missing=False):
    """公開日+1日以降のデータをチェックし、必要ならダウンロード"""
    print(f"\n=== {target_date} 以降のデータチェック ===")
    
    # ターゲット日付からチェック開始
    current_date = datetime.strptime(target_date, '%Y-%m-%d').date()
    next_date = current_date + timedelta(days=1)
    
    missing_days = []
    
    # 最大30日先までチェック
    for i in range(1, 31):
        check_date = next_date + timedelta(days=i-1)
        date_str = check_date.strftime('%Y-%m-%d')
        
        print(f"\n--- {date_str} をチェック ---")
        
        # データベースにデータがあるか確認
        db_data = get_db_data_by_date(date_str)
        db_count = len(db_data)
        
        # TDNETサイトからデータを取得
        web_date_str = check_date.strftime('%Y%m%d')
        web_data = scrape_one_day(web_date_str)
        web_count = len(web_data)
        
        print(f"データベース件数: {db_count}件")
        print(f"TDNETサイト件数: {web_count}件")
        
        if db_count == 0 and web_count > 0:
            print(f"⚠️  {date_str} のデータがデータベースにありません")
            missing_days.append(date_str)
        elif db_count == web_count:
            print(f"✅ {date_str} のデータは一致しています")
        elif db_count < web_count:
            print(f"⚠️  {date_str} はデータベースの件数が少ないです (DB:{db_count} < Web:{web_count})")
            missing_days.append(date_str)
        else:
            print(f"ℹ️  {date_str} はデータベースの件数が多いです (DB:{db_count} > Web:{web_count})")
        
        # TDNETサイトにデータがない場合はチェック終了
        if web_count == 0:
            print(f"📅 {date_str} 以降のTDNETサイトデータはありません")
            break
    
    # ダウンロード処理
    if download_missing and missing_days:
        print(f"\n📥 欠損日数分のデータをダウンロードします: {missing_days}")
        download_missing_data(missing_days)
    elif missing_days:
        print(f"\n⚠️  欠損日数: {len(missing_days)}日")
        print(f"   欠損日: {missing_days}")
        print("   --download-missing オプションでダウンロードできます")
    else:
        print(f"\n✅ {target_date} 以降のデータはすべて最新です")
    
    return missing_days


def download_missing_data(missing_days):
    """欠損日数分のデータをダウンロードしてExcel出力"""
    if not missing_days:
        print("ダウンロードするデータがありません")
        return
    
    all_rows = []
    start_date = min(missing_days)
    end_date = max(missing_days)
    
    print(f"\n📥 {start_date} から {end_date} までのデータをダウンロード中...")
    
    current_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    while current_date <= end_date_obj:
        date_str = current_date.strftime('%Y%m%d')
        if current_date.strftime('%Y-%m-%d') in missing_days:
            print(f"{date_str} 取得中...", end=" ", flush=True)
            
            day_start = time.time()
            rows = scrape_one_day(date_str)
            day_end = time.time()
            
            all_rows.extend(rows)
            print(f"({len(rows)}件, {day_end - day_start:.2f}秒)")
        else:
            print(f"{date_str} はスキップ (データベースに存在)")
        
        current_date += timedelta(days=1)
    
    if all_rows:
        # Excelファイル名を作成
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"TDNET_Missing_{start_date}_to_{end_date}_{timestamp}.xlsx"
        
        print(f"\n📊 Excel出力中: {excel_filename}")
        to_excel(all_rows, excel_filename)
        print(f"✅ ダウンロード完了: {len(all_rows)}件")
    else:
        print("❌ ダウンロードするデータがありませんでした")


def check_and_download_sequence(target_date, download_missing=False):
    """一連のチェックとダウンロード処理"""
    # 1. 指定日の差分チェック
    print(f"🔍 {target_date} の差分チェックを実行...")
    result = check_daily_diff(target_date, export_csv=False)
    
    # 2. データが一致している場合、次の日以降をチェック
    if result['is_match']:
        print(f"✅ {target_date} のデータは一致しています")
        missing_days = check_and_download_next_days(target_date, download_missing)
        return missing_days
    else:
        print(f"⚠️  {target_date} に差分があります。先に差分を解決してください")
        return []


def scrape_one_day(date_str: str) -> List[Dict]:
    """指定日の全ページ（100件単位）を走査して結合。"""
    sess = requests.Session()
    all_rows: List[Dict] = []
    page = 1
    while True:
        page_path = f"I_list_{page:03d}_{date_str}.html"
        html = fetch_page_html(sess, page_path)
        if not html:
            # 1ページ目から取れない＝該当日なし／終端
            break
        rows = parse_rows(html, date_str)
        if not rows:
            break
        all_rows.extend(rows)
        page += 1
        if page > 50:  # セーフティ
            break
    return all_rows


def to_excel(rows: List[Dict], out_path: str):
    """ハイパーリンク付きでExcelに整形出力（青白ストライプ、枠固定、H列=公開日）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TDNET"

    # ヘッダー
    ws.append(COLS)

    # データ本体
    for r in rows:
        ws.append([
            r["時刻"], r["コード"], r["会社名"],
            r["表題"], r["XBRL"], r["上場取引所"], r["更新履歴"],
            r["公開日"],  # 日付型
        ])

    # ハイパーリンク（D=表題, E=XBRL）
    for i, r in enumerate(rows, start=2):
        if r.get("表題URL"):
            c = ws.cell(row=i, column=4)
            c.hyperlink = r["表題URL"]
            c.style = "Hyperlink"
        if r.get("XBRLURL"):
            c = ws.cell(row=i, column=5)
            c.hyperlink = r["XBRLURL"]
            c.style = "Hyperlink"

    # H列（公開日）を yy/mm/dd 表示に
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=8).number_format = "yy/mm/dd"

    # テーブル（A～H）
    last_row = ws.max_row
    ref = f"A1:{get_column_letter(ws.max_column)}{last_row}"
    tbl = Table(displayName="TDNET_List", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)

    # 体裁
    widths = {"A": 8, "B": 8, "C": 18, "D": 80, "E": 8, "F": 10, "G": 12, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    wb.properties.title = "TDNET 適時開示（期間統合＋公開日）"
    wb.save(out_path)


def main():
    print(VERSION)
    ap = argparse.ArgumentParser(description="TDnet（期間指定）A～G原表＋H=公開日(yy/mm/dd)")
    ap.add_argument("--start", help="開始日 YYYYMMDD 例: 20250804")
    ap.add_argument("--end",   help="終了日 YYYYMMDD 例: 20250810")
    ap.add_argument("--out",   help="出力Excelパス（省略時: TDNET_YYYYMMDD_YYYYMMDD.xlsx）")
    ap.add_argument("--check-diff", action="store_true", help="最新データの日付で差分チェックを実行")
    ap.add_argument("--check-date", help="指定日付で差分チェックを実行 YYYY-MM-DD")
    ap.add_argument("--export-csv", action="store_true", help="差分データをCSVファイルに出力")
    ap.add_argument("--check-sequence", action="store_true", help="一連のチェックとダウンロード処理を実行")
    ap.add_argument("--download-missing", action="store_true", help="欠損データをダウンロード")
    args = ap.parse_args()

    # 一連のチェックとダウンロード処理
    if args.check_sequence:
        if args.check_diff:
            # 最新データの日付を取得
            latest_data = get_latest_db_data()
            if not latest_data:
                print("最新データの取得に失敗しました")
                return
            # datetimeオブジェクトを文字列に変換
            if hasattr(latest_data['date'], 'strftime'):
                target_date = latest_data['date'].strftime('%Y-%m-%d')
            else:
                target_date = str(latest_data['date']).split()[0]  # 文字列の場合
            print(f"最新データの日付: {target_date}")
        elif args.check_date:
            target_date = args.check_date
        else:
            print("エラー: --check-sequence には --check-diff または --check-date が必要です")
            return
        
        # 一連の処理を実行
        missing_days = check_and_download_sequence(target_date, args.download_missing)
        return

    # 差分チェックモード
    if args.check_diff or args.check_date:
        if args.check_diff:
            # 最新データの日付を取得
            latest_data = get_latest_db_data()
            if not latest_data:
                print("最新データの取得に失敗しました")
                return
            # datetimeオブジェクトを文字列に変換
            if hasattr(latest_data['date'], 'strftime'):
                target_date = latest_data['date'].strftime('%Y-%m-%d')
            else:
                target_date = str(latest_data['date']).split()[0]  # 文字列の場合
            print(f"最新データの日付: {target_date}")
        else:
            target_date = args.check_date
        
        # 差分チェック実行
        result = check_daily_diff(target_date, export_csv=args.export_csv)
        return

    # 通常のExcel出力モード
    if not args.start or not args.end:
        print("エラー: --start と --end を指定するか、--check-diff または --check-date を使用してください")
        return

    d0 = datetime.strptime(args.start, "%Y%m%d").date()
    d1 = datetime.strptime(args.end,   "%Y%m%d").date()
    if d1 < d0:
        sys.exit("エラー: --end は --start 以降の日付を指定してください。")

    out_path = args.out or f"TDNET_{d0.strftime('%Y%m%d')}_{d1.strftime('%Y%m%d')}.xlsx"

    start_time_all = time.time()  # 全体開始時間
    all_rows: List[Dict] = []
    d = d0
    while d <= d1:
        ds = d.strftime("%Y%m%d")
        print(f"{ds} 取得中...", end=" ", flush=True)
        
        day_start = time.time()  # 日次開始時間
        rows = scrape_one_day(ds)
        day_end = time.time()    # 日次終了時間
        
        all_rows.extend(rows)
        print(f"({day_end - day_start:.2f}秒)")
        d += timedelta(days=1)

    print(f"合計 {len(all_rows)} 件。Excel出力: {out_path}")
    to_excel(all_rows, out_path)
    
    end_time_all = time.time()  # 全体終了時間
    print(f"完了。 (総実行時間: {end_time_all - start_time_all:.2f}秒)")


if __name__ == "__main__":
    main()