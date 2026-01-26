#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TDnet適時開示情報.xlsm の決算期・四半期自動判定スクリプト
追加機能: ファイル使用中チェックと対話メッセージ
"""

import re
import unicodedata
import time
import os
from datetime import datetime
from calendar import monthrange
import openpyxl

# =================================================================
# 1. 設定エリア
# =================================================================
TARGET_FILE_PATH = r'\\LS720D7A9\TakashiBK\投資\TDNET\TDnet適時情報開示サービス\TDnet適時開示情報.xlsm'
START_ROW = 39649
# =================================================================

ERA_TO_YEAR = {
    '令和': 2018, '平成': 1988, '昭和': 1925, '大正': 1911, '明治': 1867
}

def normalize_text(text):
    if not text or not isinstance(text, str): return ""
    normalized = unicodedata.normalize('NFKC', text)
    return re.sub(r'\s+', ' ', normalized).strip()

def era_to_western(era_name, era_year_str):
    base_year = ERA_TO_YEAR.get(era_name)
    if base_year is None: return None
    era_year = 1 if era_year_str == '元' else int(era_year_str)
    return base_year + era_year

def extract_report_type(text):
    normalized = normalize_text(text)
    keywords = ["業績予想", "事業計画", "中期経営", "決算説明", "決算短信"]
    for kw in keywords:
        if kw in normalized: return kw
    return None

def extract_fiscal_period(text):
    normalized = normalize_text(text)
    p1 = r'(\d{4})年([1-9]|1[0-2])月(期)?'
    m1 = re.search(p1, normalized)
    if m1: return (int(m1.group(1)), int(m1.group(2)))
    p2 = r'(令和|平成|昭和|大正|明治)(元|\d+)年([1-9]|1[0-2])月(期)?'
    m2 = re.search(p2, normalized)
    if m2:
        year = era_to_western(m2.group(1), m2.group(2))
        if year: return (year, int(m2.group(3)))
    return None

def extract_quarter(text):
    normalized = normalize_text(text)
    p1 = r'([1-4])\s*Q|Q\s*([1-4])'
    m1 = re.search(p1, normalized, re.IGNORECASE)
    if m1: return f"{m1.group(1) or m1.group(2)}Q"
    p2 = r'第\s*([一二三四１２３４1-4])\s*四\s*半\s*期'
    m2 = re.search(p2, normalized)
    if m2:
        q_map = {'一':'1','二':'2','三':'3','四':'4','１':'1','２':'2','３':'3','４':'4','1':'1','2':'2','3':'3','4':'4'}
        return f"{q_map.get(m2.group(1), '4')}Q"
    if re.search(r'上半期|上期|中間期|中間', normalized): return '2Q'
    if re.search(r'下半期|下期|通期', normalized): return '4Q'
    return None

def validate_results(ws, start_row, max_row):
    print("\n=== 検収チェック開始 ===")
    error_count = 0
    for row_idx in range(start_row, min(max_row + 1, start_row + 1000)):
        if '第' in str(ws[f'D{row_idx}'].value) and ws[f'L{row_idx}'].value == '4Q':
            error_count += 1
    print(f"  チェック1: 第X四半期が4Qの行数(サンプル): {error_count}")
    print("=== 検収チェック完了 ===\n")

def is_file_writable(filepath):
    """ファイルが書き込み可能か（他のプロセスに開かれていないか）チェック"""
    if not os.path.exists(filepath):
        return True
    try:
        # 追記モードで開いてみて、すぐに閉じる
        f = open(filepath, 'a')
        f.close()
        return True
    except IOError:
        return False

def process_tdnet_file(file_path, start_row):
    start_time = time.time()
    
    # 読み込み
    print(f"ファイルを読み込み中: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
    except Exception as e:
        print(f"エラー: 読み込み失敗: {e}")
        return

    ws = wb.active
    max_row = ws.max_row
    updated_count = 0
    
    # データ処理
    for row_idx in range(start_row, max_row + 1):
        title = ws[f'D{row_idx}'].value
        if not title or not isinstance(title, str): continue
        
        # J列: 種別
        rtype = extract_report_type(title)
        if rtype: ws[f'J{row_idx}'].value = rtype
        
        # K, L列: 決算期, 四半期
        period = extract_fiscal_period(title)
        if period:
            ws[f'K{row_idx}'].value = datetime(period[0], period[1], monthrange(period[0], period[1])[1])
            ws[f'K{row_idx}'].number_format = 'yy/mm/dd'
            ws[f'L{row_idx}'].value = extract_quarter(title) or '4Q'
            updated_count += 1

    print(f"\n処理完了: {updated_count}件のデータをメモリ上にセットしました。")
    validate_results(ws, start_row, max_row)

    # 保存処理の判定
    if is_file_writable(file_path):
        try:
            wb.save(file_path)
            print(f"保存に成功しました！ (経過時間: {time.time() - start_time:.2f}秒)")
        except Exception as e:
            print(f"保存エラーが発生しました: {e}")
    else:
        # ファイルが開かれている場合のメッセージ
        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        print(f"警告: ファイルがExcel等で開かれているため、保存ができません。")
        print(f"対象: {file_path}")
        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
        
        user_input = input("このまま保存せずに処理を終了しますか？ (y/n): ").lower()
        if user_input == 'y':
            print("保存せずに終了します。データは更新されていません。")
        else:
            print("プログラムを中断しました。ファイルを閉じてから再実行してください。")

if __name__ == '__main__':
    print(f"--- TDnet 判定スクリプト (ファイルチェック機能付) ---")
    try:
        process_tdnet_file(TARGET_FILE_PATH, START_ROW)
    except Exception as e:
        print(f"予期せぬエラー: {e}")