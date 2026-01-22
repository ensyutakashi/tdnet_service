import os
import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
import time

# ================= config =================
# スクリプトがあるフォルダを作業ディレクトリにする
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

# ファイル名（パスをBASE_DIR基準で結合）
EXCEL_FILE = os.path.join(BASE_DIR, "TDnet適時開示情報.xlsm")
SHEET_NAME = '適時開示情報'

# 保存先フォルダ
PDF_FOLDER = os.path.join(BASE_DIR, "TDnet(決算短信)-PDF-随時追加分")
XBRL_FOLDER = os.path.join(BASE_DIR, "TDnet(決算短信)XBRL-随時追加分")

# 開始行 (VBAの37690行目付近。必要に応じて調整してください)
START_ROW_INDEX = 37504
MAX_WORKERS = 15  # 同時実行数
# ==========================================

def get_timestamp_msg(msg):
    return f"{msg} {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"

def download_file(url, save_path):
    try:
        if not url or not str(url).startswith('http'):
            return "失敗: URL不正"
        
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            with open(save_path, 'wb') as f:
                f.write(response.content)
            return "成功" if os.path.getsize(save_path) > 0 else "失敗: 空ファイル"
        return f"失敗: ステータス {response.status_code}"
    except Exception as e:
        return f"失敗: {str(e)}"

def is_empty(value):
    """セルが実質的に空かどうかを判定"""
    if value is None:
        return True
    if str(value).strip() == "":
        return True
    return False

def main():
    start_time = time.time() # 計測開始

    if not os.path.exists(EXCEL_FILE):
        print(f"エラー: Excelファイルが見つかりません\n{EXCEL_FILE}")
        return

    # フォルダ作成
    os.makedirs(PDF_FOLDER, exist_ok=True)
    os.makedirs(XBRL_FOLDER, exist_ok=True)

    print(f"Excelファイルを読み込んでいます: {os.path.basename(EXCEL_FILE)}")
    # keep_vba=Trueでマクロを保持、data_only=Falseで数式/リンクを保持
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False, keep_vba=True)
    ws = wb[SHEET_NAME]
    
    headers = [cell.value for cell in ws[1]]
    try:
        idx_filename = headers.index("ファイル名(連番+公開日+時刻+(種別)+決算月+4Q+コード+会社名+表題)")
        idx_pdf_res = headers.index("pdfDL")
        idx_xbrl_url = headers.index("XBRL")
        idx_xbrl_res = headers.index("xbrlDL")
        idx_pdf_url = 3 # D列(0,1,2,3)
    except ValueError as e:
        print(f"エラー: Excelのヘッダー名が一致しません。{e}")
        return

    # 指定行から最終行まで取得
    rows = list(ws.iter_rows(min_row=START_ROW_INDEX))
    total_rows = len(rows)
    print(f"{total_rows} 件の行を確認します（開始行: {START_ROW_INDEX}, 並列数: {MAX_WORKERS}）...")

    pdf_count = 0
    xbrl_count = 0

    def process_row(row):
        nonlocal pdf_count, xbrl_count
        row_num = row[0].row
        
        # PDF処理
        pdf_res_cell = row[idx_pdf_res]
        if is_empty(pdf_res_cell.value):
            url_cell = row[idx_pdf_url]
            url = url_cell.hyperlink.target if url_cell.hyperlink else url_cell.value
            fname = row[idx_filename].value
            if fname and not is_empty(url):
                if not str(fname).lower().endswith('.pdf'): fname = str(fname) + '.pdf'
                save_path = os.path.join(PDF_FOLDER, fname)
                res = download_file(url, save_path)
                pdf_res_cell.value = get_timestamp_msg(res)
                if "成功" in res: pdf_count += 1

        # XBRL処理
        xbrl_res_cell = row[idx_xbrl_res]
        if is_empty(xbrl_res_cell.value):
            url_cell = row[idx_xbrl_url]
            url = url_cell.hyperlink.target if url_cell.hyperlink else url_cell.value
            fname = row[idx_filename].value
            if fname and not is_empty(url):
                fname_zip = str(fname).replace('.pdf', '').replace('.PDF', '') + '.zip'
                save_path = os.path.join(XBRL_FOLDER, fname_zip)
                res = download_file(url, save_path)
                xbrl_res_cell.value = get_timestamp_msg(res)
                if "成功" in res: xbrl_count += 1
        
        if row_num % 100 == 0:
            print(f"進捗: {row_num} 行目を処理中...")

    # スレッドプールで並列実行
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        executor.map(process_row, rows)

    # 別名で保存
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = EXCEL_FILE.replace(".xlsm", f"_result_{timestamp}.xlsm")
    
    print("Excelファイルを保存中...")
    wb.save(output_file)
    
    # 時間計算
    end_time = time.time()
    elapsed_sec = int(end_time - start_time)
    minutes = elapsed_sec // 60
    seconds = elapsed_sec % 60

    print("\n" + "="*40)
    print(f" 処理完了！")
    print(f" 実行時間: {minutes}分 {seconds}秒")
    print(f" 結果保存先: {os.path.basename(output_file)}")
    print(f" 新規PDF取得: {pdf_count} 件")
    print(f" 新規XBRL取得: {xbrl_count} 件")
    print("="*40)

    # 保存先フォルダを開く
    os.startfile(BASE_DIR)

if __name__ == "__main__":
    main()